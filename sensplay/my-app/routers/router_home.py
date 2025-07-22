from controllers.funciones_login import *
from controllers.funciones_home import obtener_datos_salto_por_sesion
from app import app
from flask import render_template, request, flash, redirect, url_for, session, jsonify, send_file
from mysql.connector.errors import Error
from flask import Blueprint
from conexion.conexionBD import connectionBD
from urllib.parse import quote
import functools


router_home = Blueprint('router_home', __name__)
# Importando conexión a BD
from controllers.funciones_home import *

# Función centralizada para validar sesión y rol
def validar_rol(roles_permitidos=None):
    if 'conectado' not in session:
        flash('Primero debes iniciar sesión.', 'error')
        return redirect(url_for('inicio'))

    if roles_permitidos is not None:
        try:
            rol_usuario = int(session.get('rol'))
        except (ValueError, TypeError):
            rol_usuario = None
        if rol_usuario not in roles_permitidos:
            flash('No tienes permiso para acceder a esta página', 'error')
            return redirect(url_for('inicio'))
    return None

@app.route("/lista-de-pacientes", methods=['GET'])
def pacientes():
    error = validar_rol()  # ✅ Permitir acceso a todos los conectados
    if error:
        return error

    return render_template(
        'public/usuarios/lista_pacientes.html',
        resp_pacientesBD=lista_pacientesBD(),
        dataLogin=dataLoginSesion()  # aquí va el rol en el HTML
    )

@app.route('/borrar-paciente/<int:id_paciente>', methods=['POST'])
def borrar_paciente(id_paciente):
    error = validar_rol([1])
    if error:
        return error
    try:
        if eliminar_pacienteBD(id_paciente):  # función en funciones_home.py
            flash("Paciente eliminado correctamente", "success")
        else:
            flash("Error al eliminar el paciente", "danger")
    except Exception as e:
        print(f"❌ Error al eliminar paciente: {e}")
        flash("Error interno al intentar eliminar", "danger")

    return redirect(url_for('pacientes'))  # Redirige a la misma ruta de lista de pacientes




@router_home.route('/sesiones-pacientes', methods=['GET'])
def sesiones_pacientes():
    # Solo renderiza la página; la búsqueda se manejará vía JS/AJAX
    error = validar_rol()
    if error:
        return error
    return render_template(
        'public/usuarios/sesiones_pacientes.html',
        dataLogin=dataLoginSesion()
    )


# router_home.py (parte relevante)

@router_home.route('/api/buscar-paciente', methods=['POST'])
def api_buscar_paciente():
    cedula = request.json.get('cedula','').strip()
    # ... validaciones ...

    paciente = obtener_paciente_por_cedula(cedula)
    if not paciente:
        return jsonify({'error':'not_found'}), 404
    return jsonify({
        'id': paciente['id_paciente'], # Opcional si no lo usas en el frontend
        'nombres': paciente['nombres'],
        'apellidos': paciente['apellidos'],
        'edad': paciente['edad'],
        'genero': paciente['genero'], # Usamos 'genero' porque la consulta tiene 'AS genero'
        'cedula': paciente['cedula'] # <-- ESTO YA ESTÁ BIEN ASÍ
    }), 200


@router_home.route('/api/update-firebase-session', methods=['POST'])
def api_update_firebase_session():
    try:
        data = request.get_json()
        estado_sesion = data.get('estado_sesion')
        paciente_data = data.get('paciente') # Esto puede ser null si el estado es 'inactivo'

        if estado_sesion not in ["activo", "inactivo"]:
            # Esto es una validación de input, está bien que genere un error 400
            return jsonify({'error': 'Estado de sesión inválido.'}), 400
        
                # Obtener el ID del usuario logueado
        user_info = dataLoginSesion()
        id_usuario_logueado = user_info.get('id')

        if id_usuario_logueado is None:
            return jsonify({'error': 'ID de usuario no disponible. Asegúrate de estar logueado.'}), 401

        id_sesion_db = None # Inicializamos la variable para el ID de la sesión de la DB
        
                # Si el estado es "activo", primero insertar en la tabla 'sesiones'
        if estado_sesion == "activo" and paciente_data and paciente_data.get('id'):
            id_paciente = paciente_data['id']
            # Llamar a la función para registrar en la base de datos local
            # ¡Capturamos el ID de sesión generado!
            id_sesion_db = registrar_sesion_db(id_paciente, id_usuario_logueado)
            print(f"Sesión registrada en la base de datos local para paciente {id_paciente} y usuario {id_usuario_logueado}. ID de Sesión generado: {id_sesion_db}")

        # Llama a la función que interactúa con Firebase
        # Esta función, como hemos configurado, lanzará una excepción
        # si Firebase no está inicializado o hay un problema al comunicarse.
        # Llama a la función que interactúa con Firebase
        actualizar_estado_sesion_firebase(
            estado_sesion, 
            paciente_data, 
            id_sesion_actual=id_sesion_db if estado_sesion == "activo" else None # <--- ¡CORRECCIÓN AQUÍ!
        )

        # Si todo va bien, devuelve éxito
        return jsonify({'message': 'Estado de sesión de Firebase actualizado y registro de sesión en DB.'}), 200

    except Exception as e:
        print(f"Error en api_update_firebase_session: {e}")
        return jsonify({'error': f'{str(e)}'}), 500


@app.route('/lista-de-areas', methods=['GET'])
def lista_areas():
    error = validar_rol([1])
    if error:
        return error
    return render_template('public/usuarios/lista_areas.html', areas=lista_areasBD(), dataLogin=dataLoginSesion())


@app.route("/lista-de-usuarios", methods=['GET'])
def usuarios():
    error = validar_rol([1])  # Solo rol 1 puede acceder
    if error:
        return error
    return render_template(
        'public/usuarios/lista_usuarios.html',
        resp_usuariosBD=lista_usuariosBD(),
        dataLogin=dataLoginSesion(),
        areas=lista_areasBD(),
        roles=lista_rolesBD()
    )

@router_home.route('/api/obtener_sesiones_paciente', methods=['GET'])
def api_obtener_sesiones_paciente():
    """
    Endpoint de API para obtener sesiones de un paciente por cédula y fecha.
    Recibe la cédula del paciente y la fecha (YYYY-MM-DD) como parámetros de consulta (query parameters).
    """
    cedula = request.args.get('cedula')
    fecha = request.args.get('fecha') # La fecha vendrá en formato 'YYYY-MM-DD' del input type="date"

    # Validar que los parámetros no estén vacíos
    if not cedula:
        return jsonify({"registros": False, "mensaje": "La cédula es un parámetro requerido."}), 400
    if not fecha:
        return jsonify({"registros": False, "mensaje": "La fecha es un parámetro requerido."}), 400

    # Llama a la función de funciones_home para obtener los datos
    result = obtener_sesiones_por_cedula_y_fecha(cedula, fecha)
    
    # Devuelve el resultado como JSON
    return jsonify(result)


@app.route("/ImagenesComplexivo", methods=['GET'])
def imagenes():
    error = validar_rol()
    if error:
        # Aquí rediriges a 'inicioCpanel' si no está conectado
        return redirect(url_for('inicioCpanel'))
    return render_template(
        'public/usuarios/imagnesComplexivo.html',
        resp_usuariosBD=lista_usuariosBD(),
        dataLogin=dataLoginSesion(),
        areas=lista_areasBD(),
        roles=lista_rolesBD()
    )


@app.route('/borrar-usuario/<string:id>', methods=['POST'])
def borrarUsuario(id):
    error = validar_rol([1])
    if error:
        return error  # si es redirección, igual funciona

    resp = eliminarUsuario(id)
    if resp:
        flash('El Usuario fue eliminado correctamente', 'success')
    else:
        flash('No se pudo eliminar el usuario', 'error')
    
    # Redirige a la misma vista de la lista de usuarios
    return redirect(url_for('usuarios'))



@app.route('/borrar-area/<string:id_area>/', methods=['GET'])
def borrarArea(id_area):
    error = validar_rol([1])
    if error:
        return error

    resp = eliminarArea(id_area)
    if resp:
        flash('El Área fue eliminada correctamente', 'success')
    else:
        flash('Hay usuarios que pertenecen a esta área', 'error')
    return redirect(url_for('lista_areas'))


@app.route("/descargar-informe-accesos/", methods=['GET'])
def reporteBD():
    error = validar_rol()
    if error:
        return error

    try:
        ruta_archivo = generarReporteExcel()
        return send_file(
            ruta_archivo,
            as_attachment=True,
            download_name=os.path.basename(ruta_archivo),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"❌ Error al descargar el reporte: {e}")
        flash("Error al generar el reporte", "error")
        return redirect(url_for('reporteAccesos'))


@app.route("/reporte-accesos", methods=['GET'])
def reporteAccesos():
    error = validar_rol()
    if error:
        return error

    try:
        userData = dataLoginSesion()
        conexion = connectionBD()
        cursor = conexion.cursor(dictionary=True)
        cursor.execute("SELECT id, CONCAT(nombre, ' ', apellido) AS nombre_completo FROM paciente")
        pacientes = cursor.fetchall()

        cursor.execute("""
            SELECT 
                sesion.id AS id,
                sesion.fecha_inicio,
                sesion.fecha_fin,
                sesion.paciente_id,
                CONCAT(paciente.nombre, ' ', paciente.apellido) AS nombre_completo,
                terapeuta.nombre AS terapeuta,
                IF(sesion.fecha_fin IS NULL, 'Activa', 'Finalizada') AS estado
            FROM 
                sesion
            INNER JOIN paciente ON sesion.paciente_id = paciente.id
            INNER JOIN terapeuta ON sesion.terapeuta_id = terapeuta.id
        """)
        sesiones = cursor.fetchall()
        conexion.close()

        return render_template(
            'public/perfil/reportes.html',
            pacientes=pacientes,
            sesiones=sesiones,
            reportes=dataReportes(),
            lastAccess=lastAccessBD(userData.get('cedula')),
            dataLogin=userData
        )
    except Exception as e:
        print(f"❌ Error en reporteAccesos: {e}")
        flash("Error al cargar los reportes", "error")
        return redirect(url_for('inicio'))


@app.route('/crear-area', methods=['GET','POST'])
def crearArea():
    error = validar_rol([1])
    if error:
        return error

    if request.method == 'POST':
        area_name = request.form['nombre_area']
        resultado_insert = guardarArea(area_name)
        if resultado_insert:
            flash('El Area fue creada correctamente', 'success')
            return redirect(url_for('lista_areas'))
        else:
            return "Hubo un error al guardar el área."
    return render_template('public/usuarios/lista_areas')


@app.route('/actualizar-area', methods=['POST'])
def updateArea():
    error = validar_rol([1])
    if error:
        return error

    if request.method == 'POST':
        nombre_area = request.form['nombre_area']
        id_area = request.form['id_area']
        resultado_update = actualizarArea(id_area, nombre_area)
        if resultado_update:
            flash('La actualización fue correcta', 'success')
            return redirect(url_for('lista_areas'))
        else:
            return "Hubo un error al actualizar el área."
    return redirect(url_for('lista_areas'))


# === RUTAS PARA DASHBOARD PACIENTES Y SESIONES ===

@app.route('/api/pacientes', methods=['GET'])
def obtener_pacientes():
    try:
        conexion_MYSQLdb = connectionBD()
        cursor = conexion_MYSQLdb.cursor(dictionary=True)
        cursor.execute("SELECT id, nombre, apellido FROM paciente")
        pacientes = cursor.fetchall()
        conexion_MYSQLdb.close()

        resultado = [{"id": p["id"], "nombre_completo": f"{p['nombre']} {p['apellido']}"} for p in pacientes]
        return jsonify(resultado)
    except Exception as e:
        print("Error en obtener_pacientes:", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/sesiones_paciente', methods=['GET'])
def sesiones_paciente():
    try:
        nombre = request.args.get('nombre')
        if not nombre:
            return jsonify([])

        nombres = nombre.split()
        if len(nombres) < 2:
            return jsonify([])

        nombre_paciente, apellido_paciente = nombres[0], nombres[1]
        conexion_MYSQLdb = connectionBD()
        cursor = conexion_MYSQLdb.cursor(dictionary=True)

        query = """
            SELECT s.id AS sesion_id, s.fecha_inicio, s.fecha_fin, t.nombre AS terapeuta, 
                   IF(s.fecha_fin IS NULL, 'En Proceso', 'Finalizada') AS estado
            FROM sesion s
            INNER JOIN paciente p ON s.paciente_id = p.id
            INNER JOIN terapeuta t ON s.terapeuta_id = t.id
            WHERE p.nombre = %s AND p.apellido = %s
            ORDER BY s.fecha_inicio DESC
        """
        cursor.execute(query, (nombre_paciente, apellido_paciente))
        sesiones = cursor.fetchall()
        conexion_MYSQLdb.close()

        return jsonify(sesiones)
    except Exception as e:
        print("Error en sesiones_paciente:", e)
        return jsonify({"error": str(e)}), 500


@app.route('/api/sesion_detalle/<int:sesion_id>', methods=['GET'])
def sesion_detalle(sesion_id):
    try:
        conexion_MYSQLdb = connectionBD()
        cursor = conexion_MYSQLdb.cursor(dictionary=True)

        cursor.execute("""
            SELECT sensor, cantidad 
            FROM detalle_balon 
            WHERE sesion_id = %s
        """, (sesion_id,))
        balon = cursor.fetchall()

        cursor.execute("""
            SELECT total_saltos 
            FROM detalle_trampolin 
            WHERE sesion_id = %s
        """, (sesion_id,))
        trampolin = cursor.fetchone()

        conexion_MYSQLdb.close()

        resultado = {
            "balon": balon,
            "trampolin": trampolin if trampolin else {"total_saltos": 0}
        }
        return jsonify(resultado)
    except Exception as e:
        print("Error en sesion_detalle:", e)
        return jsonify({"error": str(e)}), 500



def dataSesiones():
    with connectionBD() as conexion:
        with conexion.cursor(dictionary=True) as cursor:
            cursor.execute("SELECT * FROM sesion")
            return cursor.fetchall()
def obtener_datos_sesion(id_sesion):
    """
    Consulta los datos del balón y trampolín para una sesión específica.
    Devuelve un diccionario con los datos procesados o None si no hay datos.
    """
    try:
        conexion = connectionBD()
        cursor = conexion.cursor(dictionary=True)

        # Datos del trampolín: total de saltos
        query_trampolin = """
            SELECT total_saltos
            FROM sensor_trampolin
            WHERE id_sesion = %s
        """
        cursor.execute(query_trampolin, (id_sesion,))
        trampolin_data = cursor.fetchone()

        # Datos del balón: presión por sensor
        query_balon = """
            SELECT S1, S2, S3, S4, S5, S6
            FROM sensor_balon
            WHERE id_sesion = %s
        """
        cursor.execute(query_balon, (id_sesion,))
        balon_data = cursor.fetchone()

        conexion.close()

        if not trampolin_data and not balon_data:
            return None

        return {
            'trampolin': trampolin_data,
            'balon': balon_data
        }
    except Exception as e:
        print(f"❌ Error en obtener_datos_sesion: {e}")
        return None

@router_home.route('/graficas/<int:sesion_id>', methods=['GET'])
def graficas_sesion(sesion_id):
    error = validar_rol()
    if error:
        return error

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    datos_salto = obtener_datos_salto_por_sesion(sesion_id)

    if not datos_salto or 'trampolin' not in datos_salto or 'balon' not in datos_salto:
        datos_salto = {
            "trampolin": [0, 0, 0],  # Muy leve, Normal, Brusco
            "balon": [0, 0, 0, 0, 0, 0]  # Sensor 1 a 6
        }

    return render_template(
        'public/usuarios/graficos.html',
        sesion_id=sesion_id,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        datos_salto=datos_salto,
        dataLogin=dataLoginSesion()
    )




@router_home.route('/reportes', methods=['GET'])
def reportes():
    error = validar_rol()
    if error:
        return error

    try:
        with connectionBD() as conexion_MYSQLdb:
            with conexion_MYSQLdb.cursor(dictionary=True) as cursor:
                cursor.execute("""
                    SELECT 
                        id, 
                        CONCAT(nombre, ' ', apellido) AS nombre_completo
                    FROM paciente
                """)
                pacientes = cursor.fetchall()

                cursor.execute("""
                    SELECT 
                        sesion.id AS id,
                        sesion.fecha_inicio,
                        sesion.fecha_fin,
                        sesion.paciente_id,
                        CONCAT(paciente.nombre, ' ', paciente.apellido) AS nombre_completo,
                        terapeuta.nombre AS terapeuta,
                        'Activa' AS estado
                    FROM 
                        sesion
                    INNER JOIN paciente ON sesion.paciente_id = paciente.id
                    INNER JOIN terapeuta ON sesion.terapeuta_id = terapeuta.id
                """)
                sesiones = cursor.fetchall()

        return render_template(
            'public/perfil/reportes.html',
            pacientes=pacientes,
            sesiones=sesiones,
            reportes=dataReportes(),
            lastAccess=lastAccessBD(dataLoginSesion().get('cedula')),
            dataLogin=dataLoginSesion()
        )
    except Exception as e:
        print(f"Error en /reportes: {e}")
        return "Error al cargar reportes", 500


@app.route("/descargar-reporte-paciente/<int:paciente_id>", methods=['GET'])
def descargar_reporte_paciente(paciente_id):
    error = validar_rol()
    if error:
        return error

    try:
        conexion = connectionBD()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT CONCAT(p.nombre, ' ', p.apellido) AS nombre_paciente
            FROM paciente p
            WHERE p.id = %s
        """, (paciente_id,))
        paciente = cursor.fetchone()

        cursor.execute("""
            SELECT 
                s.id AS id_sesion,
                s.fecha_inicio,
                s.fecha_fin,
                t.nombre AS terapeuta,
                IF(s.fecha_fin IS NULL, 'Activa', 'Finalizada') AS estado
            FROM sesion s
            INNER JOIN terapeuta t ON s.terapeuta_id = t.id
            WHERE s.paciente_id = %s
            ORDER BY s.fecha_inicio DESC
        """, (paciente_id,))
        sesiones = cursor.fetchall()
        conexion.close()

        if not sesiones:
            flash('No hay sesiones para este paciente', 'warning')
            return redirect(url_for('reporteAccesos'))

        import openpyxl
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sesiones"

        headers = ['ID Sesión', 'Fecha Inicio', 'Fecha Fin', 'Estado', 'Terapeuta']
        ws.append(headers)

        for sesion in sesiones:
            ws.append([
                sesion['id_sesion'],
                sesion['fecha_inicio'],
                sesion['fecha_fin'],
                sesion['estado'],
                sesion['terapeuta']
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        nombre_archivo = f"Reporte_{paciente['nombre_paciente'].replace(' ', '_')}.xlsx"

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"❌ Error al generar reporte Excel: {e}")
        flash("Error al generar el reporte Excel", "error")
        return redirect(url_for('reporteAccesos'))

@router_home.route('/formulario-paciente', methods=['GET'])
def formulario_paciente():
    generos = obtenerGeneros()
    return render_template(
        'public/usuarios/formulario-paciente.html',
        generos=generos,
        dataLogin=dataLoginSesion()  # <--- aquí
    )

@router_home.route('/guardar-paciente', methods=['POST'])
def guardar_paciente():
    try:
        cedula = request.form['cedula'].strip()
        nombres = request.form['nombres']
        apellidos = request.form['apellidos']
        edad = int(request.form['edad'])
        id_genero = int(request.form['id_genero'])

        resultado = guardarPacienteBD(cedula, nombres, apellidos, edad, id_genero)

        if resultado == "ok":
            estado = "ok"
            mensaje = "Paciente guardado correctamente"
        elif resultado == "duplicado":
            estado = "error"
            mensaje = "La cédula ya está registrada"
        else:
            estado = "error"
            mensaje = "Error al guardar paciente"

    except Exception as e:
        print("❌ Error al procesar el formulario:", e)
        estado = "error"
        mensaje = "Error interno del servidor"

    return redirect(
        url_for('router_home.formulario_paciente') +
        f"?estado={estado}&mensaje={quote(mensaje)}"
    )
